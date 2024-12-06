import openpyxl
import pandas as pd
import re
from openpyxl import load_workbook


class ExcelReader:

    str_to_task = ""

    def __init__(self, address):
        self.address = address
        ExcelReader.str_to_task = ""

    def string_handler(self): #обработчик строк, для удаления лишних символов и поиска нужных подстрок в оставшемся тексте
        symbols_to_remove = "None,()'"

        for symbol in symbols_to_remove:
            ExcelReader.str_to_task = ExcelReader.str_to_task.replace(symbol, "")


        text_list = ExcelReader.str_to_task.split()
        set_of_parts = []
        index_word = 0
        last_word = ''
        for word in text_list:
            if word == "вида" and last_word == "деталей":
                set_of_parts.append(text_list[index_word - 2])
            last_word = word
            index_word += 1

        set_of_parts = self.num_convert(set_of_parts)
        return ExcelReader.str_to_task, set_of_parts

    def reader(self):

        workbook = load_workbook(filename=self.address)
        sheet = workbook.active

        list_to_num = []


        count = 0
        sec_count = 0
        for row in sheet.iter_rows(values_only=True):
            if count == 0 or count == 9:
                ExcelReader.str_to_task += str(row)
            if count >= 5 and count < 8:
                for ind in row:
                    if sec_count > 0:
                        list_to_num.append(ind)
                    sec_count += 1
            sec_count = 0
            count += 1
        text_to_task, set_of_parts = self.string_handler()
        return list_to_num, set_of_parts, text_to_task

    def num_convert(self, set_of_parts): #метод для конвертирования чисел записанных текстом в int-овые значения
        num_dict = { 1 : "одной",
                     2 : "двух",
                     3 : "трех",
                     4 : "четырех",
                     5 : "пяти",
                     6 : "шести",
                     7 : "семи",
                     8 : "восьми",
                     9 : "девяти"}

        num_set = []

        for num in set_of_parts:
            keys = list(filter(lambda key: num_dict[key] == num, num_dict))
            num_set.append(keys[0])


        return num_set


    def xlsx_downloader(list_with_data):
        data = list_with_data
        xlsx = openpyxl.Workbook()
        sheet = xlsx.active

        sheet['A1'] = data[0]
        sheet['A2'] = data[1]
        sheet['A3'] = data[2]
    #    for row in data:
    #        sheet.append(str(row))


        xlsx.save('appending.xlsx')